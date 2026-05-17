"""
Lead service – CRUD, import, bulk operations, search.
"""

import csv
import io
import uuid
from math import ceil
from typing import Any

from sqlalchemy import delete, func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.models.campaign import Campaign, CampaignLead
from app.models.lead import Company, Contact, Lead, LeadActivity, ImportBatch, LeadScore
from app.schemas.lead import (
    LeadCreate,
    LeadDetailResponse,
    LeadResponse,
    LeadUpdate,
    CompanyResponse,
    ContactResponse,
    LeadScoreResponse,
)
from app.schemas.common import PaginatedData


class LeadService:
    def __init__(self, db: AsyncSession, tenant_id: uuid.UUID):
        self.db = db
        self.tenant_id = tenant_id

    async def list_leads(
        self,
        page: int = 1,
        page_size: int = 25,
        status: str | None = None,
        source: str | None = None,
        search: str | None = None,
        tags: list[str] | None = None,
        sort_by: str = "created_at",
        sort_dir: str = "desc",
    ) -> PaginatedData[LeadResponse]:
        query = select(Lead).where(Lead.tenant_id == self.tenant_id)

        if status:
            query = query.where(Lead.status == status)
        if source:
            query = query.where(Lead.source == source)

        # Count
        count_q = select(func.count()).select_from(query.subquery())
        total = (await self.db.execute(count_q)).scalar() or 0

        # Sort
        sort_col = getattr(Lead, sort_by, Lead.created_at)
        query = query.order_by(sort_col.desc() if sort_dir == "desc" else sort_col.asc())

        # Paginate
        query = query.offset((page - 1) * page_size).limit(page_size)
        result = await self.db.execute(query)
        leads = result.scalars().all()

        # Fetch latest score per lead in one query
        lead_ids = [l.id for l in leads]
        score_map: dict[uuid.UUID, LeadScore] = {}
        if lead_ids:
            latest_score_subq = (
                select(LeadScore.lead_id, func.max(LeadScore.created_at).label("max_ts"))
                .where(LeadScore.lead_id.in_(lead_ids))
                .group_by(LeadScore.lead_id)
                .subquery()
            )
            scores_q = await self.db.execute(
                select(LeadScore).join(
                    latest_score_subq,
                    (LeadScore.lead_id == latest_score_subq.c.lead_id)
                    & (LeadScore.created_at == latest_score_subq.c.max_ts),
                )
            )
            for s in scores_q.scalars().all():
                score_map[s.lead_id] = s

        # Fetch latest campaign per lead in one query
        campaign_map: dict[uuid.UUID, tuple[uuid.UUID, str]] = {}
        if lead_ids:
            cl_subq = (
                select(CampaignLead.lead_id, func.max(CampaignLead.created_at).label("max_ts"))
                .where(CampaignLead.lead_id.in_(lead_ids))
                .where(CampaignLead.tenant_id == self.tenant_id)
                .group_by(CampaignLead.lead_id)
                .subquery()
            )
            camp_q = await self.db.execute(
                select(CampaignLead.lead_id, Campaign.id, Campaign.name)
                .join(cl_subq, (CampaignLead.lead_id == cl_subq.c.lead_id) & (CampaignLead.created_at == cl_subq.c.max_ts))
                .join(Campaign, CampaignLead.campaign_id == Campaign.id)
            )
            for row in camp_q.all():
                campaign_map[row.lead_id] = (row.id, row.name)

        items = []
        for l in leads:
            row = LeadResponse.model_validate(l)
            sc = score_map.get(l.id)
            if sc:
                row.score_tier = sc.tier
                row.score_value = round(sc.overall_score)
            camp = campaign_map.get(l.id)
            if camp:
                row.active_campaign_id, row.active_campaign_name = camp
            items.append(row)

        return PaginatedData(
            items=items,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=ceil(total / page_size) if total > 0 else 0,
            has_next=page * page_size < total,
            has_prev=page > 1,
        )

    async def create_lead(self, data: LeadCreate) -> Lead:
        lead = Lead(
            tenant_id=self.tenant_id,
            company_id=data.company_id,
            contact_id=data.contact_id,
            source=data.source,
            tags=data.tags,
            custom_fields=data.custom_fields,
        )
        self.db.add(lead)
        await self.db.flush()

        # Log activity
        activity = LeadActivity(
            tenant_id=self.tenant_id,
            lead_id=lead.id,
            activity_type="created",
            title="Lead created",
        )
        self.db.add(activity)
        await self.db.flush()

        return lead

    async def get_lead(self, lead_id: uuid.UUID) -> LeadDetailResponse | None:
        result = await self.db.execute(
            select(Lead).where(Lead.id == lead_id, Lead.tenant_id == self.tenant_id)
        )
        lead = result.scalar_one_or_none()
        if not lead:
            return None

        return LeadDetailResponse(
            **LeadResponse.model_validate(lead).model_dump(exclude={"company", "contact", "scores"}),
            company=CompanyResponse.model_validate(lead.company) if lead.company else None,
            contact=ContactResponse.model_validate(lead.contact) if lead.contact else None,
            scores=[LeadScoreResponse.model_validate(s) for s in (lead.scores or [])],
        )

    async def update_lead(self, lead_id: uuid.UUID, data: LeadUpdate) -> Lead | None:
        result = await self.db.execute(
            select(Lead).where(Lead.id == lead_id, Lead.tenant_id == self.tenant_id)
        )
        lead = result.scalar_one_or_none()
        if not lead:
            return None

        update_data = data.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            setattr(lead, key, value)
        await self.db.flush()
        return lead

    async def delete_lead(self, lead_id: uuid.UUID) -> bool:
        result = await self.db.execute(
            delete(Lead).where(Lead.id == lead_id, Lead.tenant_id == self.tenant_id)
        )
        return result.rowcount > 0

    async def start_import(self, file, user_id: uuid.UUID) -> ImportBatch:
        content = await file.read()
        filename: str = file.filename or ""

        # Parse rows
        rows: list[dict[str, Any]] = []
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")  # handles BOM
            reader = csv.DictReader(io.StringIO(text))
            rows = [row for row in reader]
        elif filename.endswith(".xlsx"):
            try:
                import openpyxl  # pyright: ignore[reportMissingModuleSource]
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for ws_row in ws.iter_rows(min_row=2, values_only=True):
                    row = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(ws_row)}
                    rows.append(row)
            except ImportError:
                raise ValueError("openpyxl is required for XLSX imports. Install it with: pip install openpyxl")

        # Normalise column names (lowercase, strip spaces, replace spaces/dashes with _)
        def norm(s: str) -> str:
            return s.lower().strip().replace(" ", "_").replace("-", "_")

        # Column aliases — maps many possible header names → canonical field
        ALIASES: dict[str, str] = {
            # first name
            "first_name": "first_name", "firstname": "first_name", "first": "first_name",
            "given_name": "first_name",
            # last name
            "last_name": "last_name", "lastname": "last_name", "last": "last_name",
            "surname": "last_name", "family_name": "last_name",
            # email
            "email": "email", "email_address": "email", "emailaddress": "email",
            "work_email": "email", "business_email": "email",
            "customer_email": "email",
            # full contact name (will be split below)
            "customer_name": "full_name", "contact_name": "full_name",
            "full_name": "full_name", "fullname": "full_name",
            # title / job title
            "title": "title", "job_title": "title", "jobtitle": "title",
            "position": "title", "role": "title",
            # department
            "department": "department", "dept": "department", "team": "department",
            # phone
            "phone": "phone", "phone_number": "phone", "mobile": "phone",
            "telephone": "phone", "direct_phone": "phone",
            # linkedin
            "linkedin": "linkedin_url", "linkedin_url": "linkedin_url",
            "linkedin_profile": "linkedin_url",
            # location / region  ("Country Region" column maps here)
            "location": "location", "city": "location", "state": "location",
            "country": "location", "region": "location",
            "country_region": "location",
            # company name  ("Name" column in the Cvent sheet = company name)
            "name": "company_name",
            "company": "company_name", "company_name": "company_name",
            "organization": "company_name", "organisation": "company_name",
            "account": "company_name", "account_name": "company_name",
            # company domain
            "domain": "domain", "website": "domain", "company_website": "domain",
            "company_domain": "domain",
            # industry
            "industry": "industry", "vertical": "industry", "sector": "industry",
            # employee count
            "employees": "employee_count", "employee_count": "employee_count",
            "company_size": "employee_count", "headcount": "employee_count",
            # Cvent-specific fields
            "named_acct": "named_acct",
            "success_experience": "success_experience",
            # source / tags
            "source": "source", "lead_source": "source",
            "tags": "tags", "tag": "tags", "label": "tags",
        }

        batch = ImportBatch(
            tenant_id=self.tenant_id,
            file_name=filename,
            status="processing",
            total_rows=len(rows),
            created_by=user_id,
        )
        self.db.add(batch)
        await self.db.flush()

        success = 0
        errors = 0
        new_lead_ids: list[uuid.UUID] = []

        for raw_row in rows:
            try:
                # Map raw headers → canonical names
                row: dict[str, str] = {}
                for raw_key, val in raw_row.items():
                    canonical = ALIASES.get(norm(raw_key))
                    if canonical:
                        row[canonical] = str(val).strip() if val else ""

                # Skip rows with no email
                email = row.get("email", "").lower().strip()
                if not email or "@" not in email:
                    errors += 1
                    continue

                # Require at least first or last name
                first = row.get("first_name", "")
                last = row.get("last_name", "")
                if not first and not last:
                    full = row.get("full_name", "").strip()
                    parts = full.split(" ", 1)
                    first = parts[0] if parts else "Unknown"
                    last = parts[1] if len(parts) > 1 else ""

                # Upsert company
                company_name = row.get("company_name", "").strip()
                company_id = None
                if company_name:
                    existing_co = (await self.db.execute(
                        select(Company).where(
                            Company.tenant_id == self.tenant_id,
                            Company.name == company_name,
                        )
                    )).scalar_one_or_none()

                    if existing_co:
                        company_id = existing_co.id
                        # Update named_acct if provided and not already set
                        named_acct = row.get("named_acct") or None
                        if named_acct:
                            cf = dict(existing_co.custom_fields or {})
                            cf["named_acct"] = named_acct
                            existing_co.custom_fields = cf
                    else:
                        company_custom: dict[str, Any] = {}
                        if row.get("named_acct"):
                            company_custom["named_acct"] = row["named_acct"]
                        co = Company(
                            tenant_id=self.tenant_id,
                            name=company_name,
                            domain=row.get("domain") or None,
                            industry=row.get("industry") or None,
                            employee_count=int(row["employee_count"]) if row.get("employee_count", "").isdigit() else None,
                            location=row.get("location") or None,
                            custom_fields=company_custom or None,
                        )
                        self.db.add(co)
                        await self.db.flush()
                        company_id = co.id

                # Upsert contact (de-dupe by email + tenant)
                existing_ct = (await self.db.execute(
                    select(Contact).where(
                        Contact.tenant_id == self.tenant_id,
                        Contact.email == email,
                    )
                )).scalar_one_or_none()

                if existing_ct:
                    contact_id = existing_ct.id
                    # Update success_experience if provided and not already set
                    success_exp = row.get("success_experience") or None
                    if success_exp:
                        cf = dict(existing_ct.custom_fields or {})
                        cf["success_experience"] = success_exp
                        existing_ct.custom_fields = cf
                else:
                    contact_custom: dict[str, Any] = {}
                    if row.get("success_experience"):
                        contact_custom["success_experience"] = row["success_experience"]
                    contact = Contact(
                        tenant_id=self.tenant_id,
                        company_id=company_id,
                        first_name=first or "Unknown",
                        last_name=last or "",
                        email=email,
                        title=row.get("title") or None,
                        department=row.get("department") or None,
                        phone=row.get("phone") or None,
                        linkedin_url=row.get("linkedin_url") or None,
                        location=row.get("location") or None,
                        custom_fields=contact_custom or None,
                    )
                    self.db.add(contact)
                    await self.db.flush()
                    contact_id = contact.id

                # Create lead (skip if already exists for this contact)
                existing_lead = (await self.db.execute(
                    select(Lead).where(
                        Lead.tenant_id == self.tenant_id,
                        Lead.contact_id == contact_id,
                    )
                )).scalar_one_or_none()

                if not existing_lead:
                    tags_raw = row.get("tags", "")
                    tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

                    lead = Lead(
                        tenant_id=self.tenant_id,
                        company_id=company_id,
                        contact_id=contact_id,
                        source=row.get("source") or "import",
                        tags=tags,
                    )
                    self.db.add(lead)
                    await self.db.flush()
                    new_lead_ids.append(lead.id)

                success += 1

            except Exception:
                errors += 1
                continue

        batch.status = "completed"
        batch.processed_rows = len(rows)
        batch.success_rows = success
        batch.error_rows = errors
        await self.db.flush()

        return batch

    async def bulk_add_tags(self, lead_ids: list[uuid.UUID], tags: list[str]) -> int:
        result = await self.db.execute(
            select(Lead).where(
                Lead.id.in_(lead_ids), Lead.tenant_id == self.tenant_id
            )
        )
        leads = result.scalars().all()
        for lead in leads:
            existing_tags = lead.tags or []
            lead.tags = list(set(existing_tags + tags))
        await self.db.flush()
        return len(leads)

    async def bulk_assign(self, lead_ids: list[uuid.UUID], assigned_to: uuid.UUID) -> int:
        result = await self.db.execute(
            update(Lead)
            .where(Lead.id.in_(lead_ids), Lead.tenant_id == self.tenant_id)
            .values(assigned_to=assigned_to)
        )
        return result.rowcount
